# pos_app.py
# Stellar Official - Full POS (single-file)
# Features:
# - Multi-user login (roles: admin, cashier)
# - Dashboard: revenue, COGS, profit, inventory value, low stock
# - Products: add / update / delete (with cost price)
# - Sales: new sale (search by product code), cart add/remove, confirm sale
# - Invoice: generate PDF (FPDF) & Excel (in-memory) and download
# - Sales records + date-range report export
# - Settings: upload logo, change password, admin create/delete user
#
# Run:
#   pip install streamlit pandas fpdf2 openpyxl
#   streamlit run pos_app.py

import streamlit as st
import sqlite3
import pandas as pd
import os
import io
import hashlib
from datetime import datetime, date
from fpdf import FPDF
from openpyxl import Workbook, load_workbook

# ---------------- CONFIG ----------------
APP_TITLE = "Stellar Official — POS"
DB_FILE = "pos.db"
MASTER_EXCEL = "sales_master.xlsx"
DEFAULT_LOGO = "logo.png"

st.set_page_config(APP_TITLE, layout="wide")

# ---------------- HELPERS ----------------
def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()

def get_conn():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def safe_rerun():
    # streamlit rerun compatibility
    if hasattr(st, "experimental_rerun"):
        st.experimental_rerun()
    else:
        st.rerun()

# ---------------- DB INIT & SEED ----------------
def init_db():
    conn = get_conn()
    cur = conn.cursor()
    # customers
    cur.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            mobile TEXT
        )
    """)
    # products
    cur.execute("""
        CREATE TABLE IF NOT EXISTS products (
            code TEXT PRIMARY KEY,
            name TEXT,
            category TEXT,
            size TEXT,
            cost_price REAL,
            price REAL,
            stock INTEGER,
            description TEXT
        )
    """)
    # users
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT,
            role TEXT
        )
    """)
    # sales (invoice meta)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sales (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            invoice_no TEXT UNIQUE,
            user TEXT,
            customer_id INTEGER,
            total REAL,
            total_cost REAL,
            created_at TEXT
        )
    """)
    # sale items
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sale_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id INTEGER,
            product_code TEXT,
            name TEXT,
            size TEXT,
            price REAL,
            cost_price REAL,
            qty INTEGER,
            total REAL
        )
    """)
    conn.commit()

    # default admin if no users
    cur.execute("SELECT COUNT(*) as c FROM users")
    if cur.fetchone()["c"] == 0:
        cur.execute("INSERT INTO users (username,password_hash,role) VALUES (?,?,?)",
                    ("admin", hash_pw("admin123"), "admin"))
        conn.commit()
    conn.close()

def seed_products():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as c FROM products")
    if cur.fetchone()["c"] == 0:
        demo = [
            ("C001","Baby Suit - Blue","Baby","S",800.0,1200.0,10,"Soft cotton baby suit"),
            ("C002","Baby Suit - Pink","Baby","S",800.0,1200.0,8,"Pink cotton baby suit"),
            ("M001","Gents Shirt - White","Gents","M",900.0,1500.0,20,"Formal shirt white"),
            ("M002","Gents Shirt - Blue","Gents","L",950.0,1600.0,15,"Casual blue shirt"),
            ("B001","Baba Suit - Traditional","Baba","Free",1500.0,2500.0,5,"Traditional style"),
            ("P001","Gents Paint - Black","Gents","32",400.0,800.0,12,"Formal paint black"),
        ]
        for p in demo:
            cur.execute("""INSERT OR IGNORE INTO products (code,name,category,size,cost_price,price,stock,description)
                           VALUES (?,?,?,?,?,?,?,?)""", p)
        conn.commit()
    conn.close()

# ---------------- INVOICE (FPDF) ----------------
class InvoicePDF(FPDF):
    pass

def generate_invoice_bytes(invoice_no, shop_info, sale_items, total, cashier, cust_name=None, cust_mobile=None, logo_path=None):
    """
    Return PDF bytes — robust handling in case fpdf returns str/bytearray/bytes.
    sale_items: list of dicts with keys: name, size, price, qty, total
    """
    pdf = InvoicePDF(format='A4')
    pdf.add_page()
    w = pdf.w - 20

    # Logo (optional)
    if logo_path and os.path.exists(logo_path):
        try:
            pdf.image(logo_path, x=10, y=8, w=35)
        except Exception:
            pass

    pdf.set_xy(10, 20)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 8, shop_info.get("name", "Stellar Official"), ln=True, align="C")
    pdf.set_font("Arial", size=10)
    if shop_info.get("addr"):
        pdf.cell(0, 5, shop_info["addr"], ln=True, align="C")
    if shop_info.get("phone"):
        pdf.cell(0, 5, f"Phone: {shop_info['phone']}", ln=True, align="C")

    pdf.ln(6)
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 5, f"Invoice: {invoice_no}", ln=True)
    pdf.cell(0, 5, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.cell(0, 5, f"Cashier: {cashier}", ln=True)
    if cust_name:
        pdf.cell(0, 5, f"Customer: {cust_name}", ln=True)
    if cust_mobile:
        pdf.cell(0, 5, f"Mobile: {cust_mobile}", ln=True)

    pdf.ln(6)
    # Table header
    pdf.set_font("Arial", "B", 10)
    pdf.cell(10, 8, "#")
    pdf.cell(80, 8, "Item")
    pdf.cell(18, 8, "Size")
    pdf.cell(25, 8, "Price", align="R")
    pdf.cell(15, 8, "Qty", align="R")
    pdf.cell(35, 8, "Total", align="R")
    pdf.ln(8)
    pdf.set_font("Arial", size=10)
    for i, it in enumerate(sale_items, start=1):
        name = it.get("name","")[:40]
        size = str(it.get("size",""))
        price = float(it.get("price",0))
        qty = int(it.get("qty",0))
        tot = float(it.get("total",0))
        pdf.cell(10, 6, str(i))
        pdf.cell(80, 6, name)
        pdf.cell(18, 6, size)
        pdf.cell(25, 6, f"{price:.2f}", align="R")
        pdf.cell(15, 6, str(qty), align="R")
        pdf.cell(35, 6, f"{tot:.2f}", align="R")
        pdf.ln(6)

    pdf.ln(6)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, f"Grand Total: {total:.2f} PKR", ln=True, align="R")
    pdf.ln(6)
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 6, "Thank you for shopping with Stellar Official!", ln=True, align="C")

    # get output from fpdf in a robust way
    try:
        res = pdf.output(dest='S')  # fpdf2 often returns bytes or str
    except TypeError:
        # fallback: write into BytesIO using legacy API
        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        return buf.read()

    # normalize to bytes
    if isinstance(res, str):
        pdf_bytes = res.encode('latin-1')
    elif isinstance(res, bytearray):
        pdf_bytes = bytes(res)
    else:
        # assume already bytes
        pdf_bytes = res

    return pdf_bytes

def append_invoice_to_master_excel(invoice_no, df_items):
    """
    Create/update a master workbook that stores each invoice as a sheet.
    """
    if os.path.exists(MASTER_EXCEL):
        wb = load_workbook(MASTER_EXCEL)
    else:
        wb = Workbook()
        # remove default empty sheet if empty
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            wb.remove(wb.active)
    sheet_name = invoice_no if len(invoice_no) <= 31 else invoice_no[-31:]
    # ensure unique sheet name
    if sheet_name in wb.sheetnames:
        sheet_name = f"{sheet_name}_{int(datetime.now().timestamp())}"
    ws = wb.create_sheet(title=sheet_name)
    headers = list(df_items.columns)
    ws.append(["invoice_no"] + headers)
    for _, row in df_items.iterrows():
        ws.append([invoice_no] + [row[col] for col in headers])
    wb.save(MASTER_EXCEL)

# ---------------- AUTH ----------------
def check_user(username, password):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT username, role FROM users WHERE username=? AND password_hash=?", (username, hash_pw(password)))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

# ---------------- UI PAGES ----------------
def login_page():
    st.header("🔐 Login")
    c1, c2 = st.columns([3,1])
    with c1:
        username = st.text_input("Username", key="login_user")
        password = st.text_input("Password", type="password", key="login_pass")
    with c2:
        if st.button("Login"):
            user = check_user(username, password)
            if user:
                st.session_state.logged_in = True
                st.session_state.user = {"username": user["username"], "role": user["role"]}
                st.success("Login successful")
                safe_rerun()
            else:
                st.error("Invalid credentials")

def sidebar_menu():
    st.sidebar.title(APP_TITLE)
    if st.session_state.get("logged_in"):
        st.sidebar.markdown(f"**User:** {st.session_state.user.get('username')}  \n**Role:** {st.session_state.user.get('role')}")
        menu = st.sidebar.radio("Menu", ["Dashboard","New Sale","Products","Sales Record","Settings","Logout"])
    else:
        menu = st.sidebar.radio("Menu", ["Login"])
    return menu

# ---- PRODUCTS PAGE ----
def products_page():
    st.header("🛍 Products Management")
    conn = get_conn()
    cur = conn.cursor()
    tab1, tab2 = st.tabs(["Add / Update Product", "Product List & Export"])
    with tab1:
        st.subheader("Add or Update a product")
        code = st.text_input("Product Code (unique)", key="p_code")
        name = st.text_input("Name", key="p_name")
        cat = st.text_input("Category", key="p_cat")
        size = st.text_input("Size", key="p_size")
        cost_price = st.number_input("Cost Price (what you paid)", min_value=0.0, format="%.2f", key="p_cost")
        price = st.number_input("Sell Price", min_value=0.0, format="%.2f", key="p_price")
        stock = st.number_input("Stock", min_value=0, step=1, key="p_stock")
        desc = st.text_area("Description", key="p_desc")
        colA, colB = st.columns(2)
        with colA:
            if st.button("Save Product"):
                if code.strip()=="" or name.strip()=="":
                    st.warning("Provide product code and name")
                else:
                    try:
                        cur.execute("""INSERT OR REPLACE INTO products (code,name,category,size,cost_price,price,stock,description)
                                       VALUES (?,?,?,?,?,?,?,?)""",
                                    (code.strip(), name.strip(), cat.strip(), size.strip(), float(cost_price), float(price), int(stock), desc.strip()))
                        conn.commit()
                        st.success("Product saved")
                    except Exception as e:
                        st.error("Save error: " + str(e))
        with colB:
            if st.button("Delete Product"):
                if code.strip()=="":
                    st.warning("Enter product code to delete")
                else:
                    cur.execute("DELETE FROM products WHERE code=?", (code.strip(),))
                    conn.commit()
                    st.success("Deleted (if existed)")
    with tab2:
        df = pd.read_sql_query("SELECT * FROM products ORDER BY name", conn)
        st.dataframe(df)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Products")
        st.download_button("Download Products (Excel)", data=buf.getvalue(), file_name="products.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    conn.close()

# ---- NEW SALE ----
def new_sale_page():
    st.header("🛒 New Sale")
    conn = get_conn()
    cur = conn.cursor()
    prod_df = pd.read_sql_query("SELECT code,name,price,cost_price,stock,size,description FROM products ORDER BY name", conn)
    if prod_df.empty:
        st.info("No products available. Add products from Products menu.")
        conn.close()
        return

    # search by code or pick from list
    col1, col2 = st.columns([2,3])
    with col1:
        search_code = st.text_input("Search by Product Code (e.g., C001)", key="search_code")
        if st.button("Find"):
            if search_code.strip():
                found = prod_df[prod_df['code'].str.upper() == search_code.strip().upper()]
                if found.empty:
                    st.warning("Product code not found")
                else:
                    row = found.iloc[0]
                    st.session_state.selected_product = dict(row)
                    safe_rerun()
    with col2:
        select = st.selectbox("Or choose product from list", options=[f"{r['code']} - {r['name']} (Stock:{r['stock']})" for _,r in prod_df.iterrows()])
        if st.button("Select from list"):
            pid = select.split(" - ")[0]
            row = prod_df[prod_df['code'] == pid].iloc[0]
            st.session_state.selected_product = dict(row)
            safe_rerun()

    # show selected
    selected = st.session_state.get("selected_product")
    if not selected:
        st.info("Select a product or search by code")
        conn.close()
        return

    st.markdown(f"**{selected['name']}** | Price: {selected['price']} | Stock: {selected['stock']}")
    st.write(selected.get("description",""))
    qty = st.number_input("Qty to add", min_value=1, value=1, step=1, key="add_qty")
    if st.button("Add to Cart"):
        if qty > selected['stock']:
            st.error("Not enough stock")
        else:
            if "cart" not in st.session_state:
                st.session_state.cart = []
            found=False
            for it in st.session_state.cart:
                if it["product_code"] == selected['code']:
                    it["qty"] += qty
                    it["total"] = it["price"] * it["qty"]
                    found=True
                    break
            if not found:
                st.session_state.cart.append({
                    "product_code": selected['code'],
                    "name": selected['name'],
                    "size": selected.get('size',''),
                    "price": float(selected['price']),
                    "cost_price": float(selected['cost_price']),
                    "qty": int(qty),
                    "total": float(selected['price']) * int(qty)
                })
            st.success("Added to cart")
            safe_rerun()

    st.subheader("Cart")
    if "cart" not in st.session_state or not st.session_state.cart:
        st.info("Cart is empty")
        conn.close()
        return

    dfc = pd.DataFrame(st.session_state.cart)
    dfc.index += 1
    st.dataframe(dfc[["product_code","name","size","price","qty","total"]], use_container_width=True)
    subtotal = float(dfc["total"].sum())
    st.markdown(f"### Subtotal: **{subtotal:.2f} PKR**")

    # remove / clear
    colA, colB = st.columns(2)
    with colA:
        rem_code = st.text_input("Remove by Product Code (leave blank)", key="rem_code")
        if st.button("Remove Item"):
            if rem_code.strip():
                st.session_state.cart = [it for it in st.session_state.cart if it["product_code"]!=rem_code.strip()]
                st.success("Removed (if existed)")
                safe_rerun()
            else:
                st.warning("Enter product code to remove")
    with colB:
        if st.button("Clear Cart"):
            st.session_state.cart = []
            safe_rerun()

    st.markdown("---")
    st.subheader("Customer & Payment")
    cust_name = st.text_input("Customer Name (optional)", key="cust_name")
    cust_mobile = st.text_input("Customer Mobile (optional)", key="cust_mobile")
    discount = st.number_input("Discount (PKR)", min_value=0.0, value=0.0, format="%.2f", key="discount")
    grand_total = max(0.0, subtotal - float(discount))
    st.markdown(f"### Grand Total: **{grand_total:.2f} PKR**")

    if st.button("Confirm & Generate Invoice"):
        # re-check stock
        ok=True
        for it in st.session_state.cart:
            cur.execute("SELECT stock FROM products WHERE code=?", (it['product_code'],))
            r = cur.fetchone()
            if not r or r['stock'] < it['qty']:
                ok=False
                st.error(f"Insufficient stock for {it['product_code']}")
                break
        if not ok:
            conn.close()
            return

        invoice_no = f"INV{datetime.now().strftime('%Y%m%d%H%M%S')}"
        user = st.session_state.user['username'] if st.session_state.get("user") else "unknown"

        # save customer
        cust_id = None
        if cust_name or cust_mobile:
            cur.execute("INSERT INTO customers (name,mobile) VALUES (?,?)", (cust_name.strip() if cust_name else "", cust_mobile.strip() if cust_mobile else ""))
            cust_id = cur.lastrowid

        created_at = datetime.now().isoformat()
        total_cost = sum([it['cost_price']*it['qty'] for it in st.session_state.cart])
        cur.execute("INSERT INTO sales (invoice_no,user,customer_id,total,total_cost,created_at) VALUES (?,?,?,?,?,?)",
                    (invoice_no, user, cust_id, grand_total, total_cost, created_at))
        sale_id = cur.lastrowid

        for it in st.session_state.cart:
            cur.execute("""INSERT INTO sale_items (sale_id,product_code,name,size,price,cost_price,qty,total)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (sale_id, it['product_code'], it['name'], it['size'], it['price'], it['cost_price'], it['qty'], it['total']))
            cur.execute("UPDATE products SET stock = stock - ? WHERE code=?", (it['qty'], it['product_code']))
        conn.commit()

        # prepare PDF bytes & Excel bytes (in-memory)
        shop_info = {"name":"Stellar Official", "addr":"Your Shop Address", "phone":"0000-000000"}
        logo_path = DEFAULT_LOGO if os.path.exists(DEFAULT_LOGO) else None

        # IMPORTANT: create a copy of sale items BEFORE we clear anything (we won't clear here,
        # but copy is safer to ensure PDF uses exact sold items)
        sale_items_for_invoice = [dict(it) for it in st.session_state.cart]

        try:
            pdf_bytes = generate_invoice_bytes(invoice_no, shop_info, sale_items_for_invoice, grand_total, user, cust_name=cust_name, cust_mobile=cust_mobile, logo_path=logo_path)
        except Exception as e:
            pdf_bytes = None
            st.error("PDF generation error: " + str(e))

        df_items = pd.DataFrame(sale_items_for_invoice).copy()
        df_items.insert(0, "invoice_no", invoice_no)
        excel_buf = io.BytesIO()
        with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
            df_items.to_excel(writer, index=False, sheet_name="Items")
            summary = pd.DataFrame({
                "Invoice": [invoice_no],
                "Customer": [cust_name],
                "Mobile": [cust_mobile],
                "Total": [grand_total]
            })
            summary.to_excel(writer, index=False, sheet_name="Summary")
        excel_bytes = excel_buf.getvalue()

        # append to master excel
        try:
            append_invoice_to_master_excel(invoice_no, df_items)
        except Exception:
            pass

        # store in session for downloads — normalize types to bytes
        st.session_state.last_invoice_no = invoice_no
        if pdf_bytes is not None:
            # ensure bytes type
            if isinstance(pdf_bytes, bytearray):
                st.session_state.last_pdf_bytes = bytes(pdf_bytes)
            else:
                st.session_state.last_pdf_bytes = pdf_bytes
        else:
            st.session_state.last_pdf_bytes = None
        st.session_state.last_excel_bytes = excel_bytes
        st.session_state.invoice_ready = True
        st.success(f"Invoice generated: {invoice_no}")

        # IMPORTANT CHANGE: DO NOT clear cart here and DO NOT force rerun.
        # Cart will be cleared only when user presses the explicit "New Sale" button below.

    # Download area if invoice_ready
    if st.session_state.get("invoice_ready"):
        st.markdown("---")
        st.markdown("### Invoice Ready — Downloads")
        if st.session_state.get("last_pdf_bytes"):
            st.download_button("⬇️ Download Invoice (PDF)", data=st.session_state["last_pdf_bytes"],
                               file_name=f"{st.session_state['last_invoice_no']}.pdf", mime="application/pdf")
        else:
            st.info("PDF not available (generation failed). You can download Excel below if available.")
        if st.session_state.get("last_excel_bytes"):
            st.download_button("⬇️ Download Invoice (Excel)", data=st.session_state["last_excel_bytes"],
                               file_name=f"{st.session_state['last_invoice_no']}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        if os.path.exists(MASTER_EXCEL):
            with open(MASTER_EXCEL, "rb") as f:
                st.download_button("⬇️ Download Master Sales Excel (All Invoices)", data=f.read(), file_name=MASTER_EXCEL,
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        # Provide two actions:
        # 1) Done (just clear invoice-ready downloads)
        # 2) New Sale (clear invoice & cart and start fresh)
        colL, colR = st.columns([3,1])
        with colL:
            if st.button("🧾 Done (Clear Invoice Ready)"):
                st.session_state.invoice_ready = False
                st.session_state.last_pdf_bytes = None
                st.session_state.last_excel_bytes = None
                st.session_state.last_invoice_no = None
                safe_rerun()
        with colR:
            if st.button("➕ New Sale"):
                # clear everything related to invoice and cart so user can start a fresh sale
                st.session_state.invoice_ready = False
                st.session_state.last_pdf_bytes = None
                st.session_state.last_excel_bytes = None
                st.session_state.last_invoice_no = None
                st.session_state.cart = []
                safe_rerun()

    conn.close()

# ---- SALES RECORD PAGE ----
def sales_record_page():
    st.header("📑 Sales Records & Reports")
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT s.id, s.invoice_no, s.user, s.total, s.total_cost, s.created_at, c.name as customer, c.mobile as mobile
        FROM sales s LEFT JOIN customers c ON s.customer_id = c.id
        ORDER BY s.created_at DESC
    """, conn)
    st.dataframe(df)

    st.subheader("Generate date-range report")
    col1, col2 = st.columns(2)
    with col1:
        d1 = st.date_input("From", value=date.today())
    with col2:
        d2 = st.date_input("To", value=date.today())

    if st.button("Generate Report for Range"):
        q = """SELECT s.invoice_no, s.user, s.total, s.total_cost, s.created_at, c.name as customer, c.mobile as mobile
               FROM sales s LEFT JOIN customers c ON s.customer_id = c.id
               WHERE DATE(s.created_at) BETWEEN ? AND ? ORDER BY s.created_at"""
        dfr = pd.read_sql_query(q, conn, params=(d1.isoformat(), d2.isoformat()))
        if dfr.empty:
            st.info("No sales in this range")
        else:
            st.dataframe(dfr)
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                dfr.to_excel(writer, index=False, sheet_name="RangeSales")
            st.download_button("Download Range Report (Excel)", data=buf.getvalue(),
                               file_name=f"sales_report_{d1}_{d2}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    inv = st.text_input("Enter Invoice No to view items", key="show_inv")
    if st.button("Show Invoice Items"):
        if inv.strip()=="":
            st.warning("Enter invoice number")
        else:
            items = pd.read_sql_query("SELECT * FROM sale_items WHERE sale_id = (SELECT id FROM sales WHERE invoice_no=?)", conn, params=(inv.strip(),))
            if items.empty:
                st.info("No items found for that invoice")
            else:
                st.dataframe(items)
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    items.to_excel(writer, index=False, sheet_name="Items")
                st.download_button("Download Items Excel", data=buf.getvalue(),
                                   file_name=f"{inv}_items.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if st.button("Export All Sales to Excel"):
        full = pd.read_sql_query("SELECT * FROM sales ORDER BY created_at DESC", conn)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            full.to_excel(writer, index=False, sheet_name="AllSales")
        st.download_button("⬇️ Download All Sales Excel", data=buf.getvalue(), file_name=f"all_sales_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    conn.close()

# ---- DASHBOARD ----
def dashboard_page():
    st.header("📊 Dashboard")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as c FROM sales")
    total_sales = cur.fetchone()["c"]
    cur.execute("SELECT SUM(total) as s FROM sales")
    revenue = cur.fetchone()["s"] or 0.0
    cur.execute("SELECT SUM(total_cost) as tc FROM sales")
    total_cost_sold = cur.fetchone()["tc"] or 0.0
    profit = revenue - total_cost_sold
    cur.execute("SELECT COUNT(*) as c FROM products")
    total_products = cur.fetchone()["c"]
    cur.execute("SELECT SUM(stock) as qty FROM products")
    total_stock = cur.fetchone()["qty"] or 0
    cur.execute("SELECT SUM(stock * cost_price) as inv FROM products")
    inv_value = cur.fetchone()["inv"] or 0.0

    st.metric("Total Sales (invoices)", total_sales)
    st.metric("Revenue (PKR)", f"{revenue:.2f}")
    st.metric("COGS (PKR)", f"{total_cost_sold:.2f}")
    st.metric("Profit (PKR)", f"{profit:.2f}")
    st.metric("Inventory Value (cost)", f"{inv_value:.2f}")
    st.metric("Total Products", total_products)
    st.metric("Total Stock Units", total_stock)

    st.subheader("Low Stock (<=5)")
    low = pd.read_sql_query("SELECT code,name,stock FROM products WHERE stock <= 5 ORDER BY stock ASC", conn)
    st.dataframe(low)

    st.subheader("Top Selling Products")
    top = pd.read_sql_query("""
        SELECT si.product_code, si.name, SUM(qty) as sold_qty
        FROM sale_items si GROUP BY si.product_code ORDER BY sold_qty DESC LIMIT 10
    """, conn)
    st.dataframe(top)
    conn.close()

# ---- SETTINGS ----
def settings_page():
    st.header("⚙️ Settings")
    st.subheader("Branding / Logo")
    st.write("Upload logo (PNG/JPG). If none, invoice will use shop name.")
    uploaded = st.file_uploader("Upload logo", type=["png","jpg","jpeg"])
    if uploaded:
        path = os.path.join(os.getcwd(), DEFAULT_LOGO)
        with open(path, "wb") as f:
            f.write(uploaded.getbuffer())
        st.success("Logo uploaded as logo.png")

    conn = get_conn()
    cur = conn.cursor()
    st.subheader("User & Password")
    if st.session_state.get("user"):
        st.write(f"Logged in as: **{st.session_state.user.get('username')}** (role: {st.session_state.user.get('role')})")
        old = st.text_input("Current Password", type="password", key="pw_old")
        newpw = st.text_input("New Password", type="password", key="pw_new")
        if st.button("Change My Password"):
            if old.strip()=="" or newpw.strip()=="":
                st.warning("Enter current and new password")
            else:
                cur.execute("SELECT * FROM users WHERE username=? AND password_hash=?", (st.session_state.user.get("username"), hash_pw(old)))
                if cur.fetchone():
                    cur.execute("UPDATE users SET password_hash=? WHERE username=?", (hash_pw(newpw), st.session_state.user.get("username")))
                    conn.commit()
                    st.success("Password changed. Please login again.")
                    st.session_state.logged_in = False
                    st.session_state.user = None
                    safe_rerun()
                else:
                    st.error("Current password incorrect")
    else:
        st.info("Login to change password")

    # Admin-only user management
    if st.session_state.get("user") and st.session_state.user.get("role") == "admin":
        st.markdown("---")
        st.subheader("Admin: Create / Delete User")
        new_un = st.text_input("New username", key="new_un")
        new_pw = st.text_input("New password", type="password", key="new_pw")
        role = st.selectbox("Role", ["cashier","admin"], key="new_role")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Create User"):
                if new_un.strip()=="" or new_pw.strip()=="":
                    st.warning("Provide username & password")
                else:
                    try:
                        cur.execute("INSERT INTO users (username,password_hash,role) VALUES (?,?,?)", (new_un.strip(), hash_pw(new_pw.strip()), role))
                        conn.commit()
                        st.success("User created")
                    except Exception as e:
                        st.error(str(e))
        with col2:
            del_un = st.text_input("Delete username", key="del_un")
            if st.button("Delete User"):
                if del_un.strip()=="":
                    st.warning("Enter username to delete")
                else:
                    if del_un.strip() == st.session_state.user.get("username"):
                        st.warning("You cannot delete yourself while logged in")
                    else:
                        cur.execute("DELETE FROM users WHERE username=?", (del_un.strip(),))
                        conn.commit()
                        st.success("Deleted if existed")
    else:
        st.info("Login as admin to manage users.")
    conn.close()

# ---------------- MAIN ----------------
def main():
    init_db()
    seed_products()
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "user" not in st.session_state:
        st.session_state.user = None
    if "cart" not in st.session_state:
        st.session_state.cart = []
    if "invoice_ready" not in st.session_state:
        st.session_state.invoice_ready = False

    st.title(APP_TITLE)
    menu = sidebar_menu()

    if menu == "Login":
        login_page()
        return

    if menu == "Logout":
        if st.button("Confirm Logout"):
            st.session_state.logged_in = False
            st.session_state.user = None
            safe_rerun()
        return

    if not st.session_state.get("logged_in"):
        st.warning("Please login from the Login menu")
        login_page()
        return

    # Route pages
    if menu == "Dashboard":
        dashboard_page()
    elif menu == "Products":
        products_page()
    elif menu == "New Sale":
        new_sale_page()
    elif menu == "Sales Record":
        sales_record_page()
    elif menu == "Settings":
        settings_page()

if __name__ == "__main__":
    main()